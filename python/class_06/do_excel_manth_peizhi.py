# 1：利用1229课堂上学习的超继承以及Excel，来给你们的数学类写20条用例（加法10条，减法10条）
# 2：要利用类与对象的思想，写一个类把数据从Excel里面读取出来 （类与对象的思想）
# 3：数据以及期望值要存到Excel里面，然后用例的执行结果要写回到Excel里面
from openpyxl import load_workbook  # 读写   导入openpyxl模块，可以进行excel文件的读写
from basic.python.class_06.peizhi_file import ReadConfig


class Cases:
    # 用来存储我们的测试数据
    def __init__(self):
        self.id = None
        self.title = None
        self.a = None
        self.b = None
        self.excepted = None


class DoExcel:
    # 老获取excel里面对应表单数据的一个类
    def __init__(self, file_name, sheet_name):

        self.file_name = file_name  # 工作簿
        self.sheet_name = sheet_name  # 表单名

    def get_data(self):  # 获取数据
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        button = ReadConfig('case.conf').get_value('CASE', 'button')
        if button == 'all':
            case = []  # 所有的数据都存在大列表里面
            for i in range(2, sheet.max_row + 1):  # 行的范围从第二行开始
                cases = Cases()  # 每一行数据存在对象Cases()里面
                cases.id = sheet.cell(row=i, column=1).value  # 存的是id
                cases.title = sheet.cell(row=i, column=2).value  # 存的是title
                cases.a = sheet.cell(row=i, column=3).value  # 存的是a
                cases.b = sheet.cell(row=i, column=4).value  # 存的是b
                cases.excepted = sheet.cell(row=i, column=5).value  # 存的是excepted
                case.append(cases)
            return case

        else:
            case = []
            for i in eval(button):
                cases = Cases()  # 每一行数据存在对象Cases()里面
                cases.id = sheet.cell(row=i + 1, column=1).value  # 存的是id
                cases.title = sheet.cell(row=i + 1, column=2).value  # 存的是title
                cases.a = sheet.cell(row=i + 1, column=3).value  # 存的是a
                cases.b = sheet.cell(row=i + 1, column=4).value  # 存的是b
                cases.excepted = sheet.cell(row=i + 1, column=5).value  # 存的是excepted
                case.append(cases)
            return case

    def write_back(self, row, column, value):  # 写回数据
        # row 行 column列  value我们要写回的值
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        sheet.cell(row, column).value = value  # 写入值到单元格里面
        wb.save(self.file_name)  # 要记得保存，同时excel要记得关闭状态


if __name__ == '__main__':
    test_data = DoExcel('python13.xlsx', 'add').get_data()
    print(test_data)
