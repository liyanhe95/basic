# 1：利用1229课堂上学习的超继承以及Excel，来给你们的数学类写20条用例（加法10条，减法10条）
# 2：要利用类与对象的思想，写一个类把数据从Excel里面读取出来 （类与对象的思想）
# 3：数据以及期望值要存到Excel里面，然后用例的执行结果要写回到Excel里面
from openpyxl import load_workbook#读写   导入openpyxl模块，可以进行excel文件的读写
# class Case:    #定义专门存储我们的测试数据的Cases类
#     def __init__(self):         #初始化方法
#         self.id=None
#         self.title=None
#         self.a=None
#         self.b=None
#         self.expected=None
#
# class DoExcel:     #来获取表单对应数据的一个类
#     def __init__(self,file_name):
#         try:   #处理文件异常
#             self.file_name = file_name
#             self.workbook = load_workbook(filename=self.file_name)
#         except FileNotFoundError as e:
#             print('{0} not found, please check file path'.format(file_name))
#             raise e
#
#     def get_cases(self,sheet_name):   #获取数据  （对象方法）
#         self.sheet_name = sheet_name
#         sheet = self.workbook[sheet_name]   #获取 sheet
#         max_row = sheet .max_row   #获取sheet最大的行数
#         #用类与对象的思想
#         cases = []   #列表
#         for i in range(2, max_row + 1):  # 行的范围从第二行开始
#             case=Case()  #实例化一个case对象，用例存放数据
#             case.id = sheet.cell(row=i,column=1).value  #存的是case_id
#             case.title = sheet.cell(row=i, column=2).value  #存的是title
#             case.url = sheet.cell(row=i, column=3).value  #存的是a
#             case.data= sheet.cell(row=i, column=4).value  #存的是b
#             case.expected = sheet.cell(row=i, column=5).value  #存的是expected
#             # if type(case.expected) == int:
#             #     case.expected = str(case.expected)
#             # cases.append(case)  #读取完毕之后把每一行的数据的值存到这个test_data这个大列表里面
#         return cases
#
#
#     def write_result(self,row,actual,result):     #定义写回数据的对象方法
#         #row 行     colnum  列   value  要写回的值
#         sheet = self.workbook[self.sheet_name]#获取sheet
#         sheet.cell(row,6).value = actual  #写入实际结果
#         sheet.cell(row,7).value = result  #写入执行结果，PASS，FAIL
#         self.workbook.save(filename=self.file_name)
class Cases:
    #用来存储我们的测试数据
    def __init__(self):
        self.id = None
        self.title = None
        self.a = None
        self.b = None
        self.excepted = None

class DoExcel:
    #老获取excel里面对应表单数据的一个类
    def __init__(self,file_name,sheet_name):

        self.file_name = file_name  # 工作簿
        self.sheet_name = sheet_name  # 表单名

    def get_data(self):#获取数据
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]

        case = []#所有的数据都存在大列表里面
        for i in range(2,sheet.max_row+1): #行的范围从第二行开始
            cases = Cases() #每一行数据存在对象Cases()里面
            cases.id = sheet.cell(row = i, column = 1).value  #存的是id
            cases.title = sheet.cell(row = i, column = 2).value#存的是title
            cases.a = sheet.cell(row = i, column = 3).value#存的是a
            cases.b = sheet.cell(row = i, column = 4).value#存的是b
            cases.excepted = sheet.cell(row = i, column = 5).value#存的是excepted
            case.append(cases)
        return case
    def write_back(self,row,column,value):#写回数据
        #row 行 column列  value我们要写回的值
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        sheet.cell(row,column).value = value #写入值到单元格里面
        wb.save(self.file_name)#要记得保存，同时excel要记得关闭状态




if __name__ == '__main__':
    test_data=DoExcel('python13.xlsx','add').get_data()
    print(test_data)