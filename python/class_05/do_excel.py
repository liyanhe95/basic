#openpyxl
#pip install openpyxl  同时支持excel的读写
#xlrd xlwt 分开读写
#只支持这个格式:xlsx
#新建文件的时候
#1.直接在桌面上新建.xlsx结尾的文件
from openpyxl import load_workbook  #支持读写
from openpyxl import Workbook #创建文件/新建文件
# Workbook 是一个对象
# 2.
# wb = Workbook() #实例化一个对象
# wb.save('python13.xlsx')  #必须要保存，否则不能创建成功
# load_workbook 是一个函数  来进行excel文件的读写
#打开excel
wb = load_workbook('python13.xlsx')
#定位表单
# sheet = wb.get_sheet_by_name('test_data')#方法一  已经过时
sheet = wb['test_data'] #方法二
#定位单元格--取值
for i in range(1,sheet.max_row+1):
    for j in range(1,sheet.max_column+1):
        res = sheet.cell(row = i,column = j).value
        print('获取到的值是:{}'.format(res))
        print('获取到的值的类型是：{}'.format(type(res)))

#如果excel里面的无值，读取出来的就是None
#定位单元格--写值
# sheet.cell(row=5,column=5,value=9)#方法1
# sheet.cell(row=6,column=6).value=10 #方法2
# wb.save('python13.xlsx')

#字符串 -- str
#列表---字符串 如果想变成原始类型 就利用eval函数 进行转换、
#字典---字符串
#float--float
#int--int

#获取最大行数和最大列数
# print("最大行：",sheet.max_row)
# print("最大列：",sheet.max_column)
